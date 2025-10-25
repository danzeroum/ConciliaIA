"""Unit tests for the export service."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from unittest.mock import AsyncMock, Mock
import pytest
from openpyxl import load_workbook

from src.application.services.export_service import ExportService
from src.domain.entities import ReconciliationMatch, Sale
from src.domain.entities.reconciliation_match import MatchType
from src.domain.value_objects import Money


def _make_sale(**kwargs) -> Sale:
    defaults = {
        "id": "sale-1",
        "tenant_id": "tenant-1",
        "nsu": "NSU123",
        "amount": Money(Decimal("150.50")),
        "date": date(2024, 1, 10),
        "payment_method": "credit",
        "authorization_code": "AUTH123",
        "installments": 2,
        "created_at": datetime(2024, 1, 11, 12, 0, 0),
    }
    defaults.update(kwargs)
    return Sale(**defaults)


@pytest.mark.asyncio
async def test_export_sales_to_excel_generates_workbook() -> None:
    sale_repo = Mock()
    match_repo = Mock()
    transaction_repo = Mock()
    divergence_repo = Mock()
    settlement_repo = Mock()

    sale_repo.find_by_date_range = AsyncMock(return_value=[_make_sale()])
    match_repo.find_by_date_range = AsyncMock(
        return_value=[
            ReconciliationMatch(
                id="match-1",
                tenant_id="tenant-1",
                sale_id="sale-1",
                transaction_id="txn-1",
                match_type=MatchType.EXACT,
                confidence=Decimal("0.98"),
            )
        ]
    )

    settlement_repo.find_by_period = AsyncMock(return_value=[])

    service = ExportService(
        sale_repo=sale_repo,
        transaction_repo=transaction_repo,
        match_repo=match_repo,
        divergence_repo=divergence_repo,
        settlement_repo=settlement_repo,
    )

    output = await service.export_sales_to_excel("tenant-1", date(2024, 1, 1), date(2024, 1, 31))

    assert isinstance(output, BytesIO)
    output.seek(0)
    workbook = load_workbook(output, data_only=True)

    sheet = workbook["Vendas"]
    headers = [cell.value for cell in sheet[4]]
    assert "NSU" in headers
    assert "Status" in headers

    status_column_index = headers.index("Status") + 1
    status_value = sheet.cell(row=5, column=status_column_index).value
    assert status_value == "Reconciliado"

    summary_sheet = workbook["Resumo"]
    summary_titles = [cell.value for cell in summary_sheet[1]]
    assert "Resumo Estatístico" in summary_titles


@pytest.mark.asyncio
async def test_export_reports_use_report_service() -> None:
    sale_repo = Mock()
    match_repo = Mock()
    transaction_repo = Mock()
    divergence_repo = Mock()

    report_service = Mock()
    report_service.generate_accuracy_report = AsyncMock(
        return_value={
            "overall_accuracy": 97.5,
            "total_sales": 10,
            "total_matches": 9,
            "total_divergences": 1,
            "trend": [
                {"date": "2024-01-01", "accuracy": 95.0, "matches": 3, "total_sales": 3}
            ],
        }
    )
    report_service.generate_divergence_analysis = AsyncMock(
        return_value={
            "total_divergences": 2,
            "total_amount_at_risk": 345.67,
            "resolution_rate": 80.0,
            "avg_resolution_time_hours": 12.5,
            "by_type": [{"type": "missing_transaction", "count": 2, "total_amount": 200.0, "percentage": 50.0}],
            "by_severity": [{"severity": "high", "count": 1, "total_amount": 150.0}],
        }
    )

    settlement_repo = Mock()
    settlement_repo.find_by_period = AsyncMock(return_value=[])

    service = ExportService(
        sale_repo=sale_repo,
        transaction_repo=transaction_repo,
        match_repo=match_repo,
        divergence_repo=divergence_repo,
        settlement_repo=settlement_repo,
        report_service=report_service,
    )

    accuracy_output = await service.export_accuracy_report_to_excel(
        "tenant-1", date(2024, 1, 1), date(2024, 1, 31)
    )
    divergence_output = await service.export_divergence_report_to_excel(
        "tenant-1", date(2024, 1, 1), date(2024, 1, 31)
    )

    assert isinstance(accuracy_output, BytesIO)
    accuracy_output.seek(0)
    accuracy_workbook = load_workbook(accuracy_output, data_only=True)
    summary = accuracy_workbook["Resumo Accuracy"]
    summary_labels = [row[0] for row in summary.iter_rows(values_only=True)]
    assert "Accuracy Geral" in summary_labels

    assert isinstance(divergence_output, BytesIO)
    divergence_output.seek(0)
    divergence_workbook = load_workbook(divergence_output, data_only=True)
    sheet = divergence_workbook["Análise Divergências"]
    divergence_labels = [row[0] for row in sheet.iter_rows(values_only=True)]
    assert "Total de Divergências" in divergence_labels

    report_service.generate_accuracy_report.assert_awaited()
    report_service.generate_divergence_analysis.assert_awaited()
