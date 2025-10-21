"""Service responsible for exporting data sets to spreadsheet formats."""

from __future__ import annotations

import io
from datetime import date
from typing import Iterable

import pandas as pd
import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from .report_service import ReportService
from src.domain.entities import Sale
from src.infrastructure.persistence.repositories import (
    DivergenceRepository,
    MatchRepository,
    SaleRepository,
    TransactionRepository,
)

logger = structlog.get_logger(__name__)


class ExportService:
    """Generate spreadsheet exports for sales and analytical reports."""

    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    def __init__(
        self,
        sale_repo: SaleRepository,
        transaction_repo: TransactionRepository,
        match_repo: MatchRepository,
        divergence_repo: DivergenceRepository,
        *,
        report_service: ReportService | None = None,
    ) -> None:
        self.sale_repo = sale_repo
        self.transaction_repo = transaction_repo
        self.match_repo = match_repo
        self.divergence_repo = divergence_repo
        self._report_service = report_service

    async def export_sales_to_excel(
        self,
        tenant_id: str,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> io.BytesIO:
        """Create an Excel export with the sales registered by the tenant."""

        period_start = start_date or date(date.today().year - 1, 1, 1)
        period_end = end_date or date.today()

        logger.info(
            "exporting_sales_to_excel",
            tenant_id=tenant_id,
            start_date=period_start.isoformat(),
            end_date=period_end.isoformat(),
        )

        sales = await self.sale_repo.find_by_date_range(
            tenant_id=tenant_id,
            start_date=period_start,
            end_date=period_end,
        )
        matches = await self.match_repo.find_by_date_range(
            tenant_id=tenant_id,
            start_date=period_start,
            end_date=period_end,
        )

        matched_sales = {match.sale_id for match in matches}
        sales_data = [self._serialize_sale(sale, matched_sales) for sale in sales]

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Vendas"

        worksheet.append(["Relatório de Vendas - ConciliaAI"])
        worksheet.append(
            [
                f"Período: {period_start.isoformat()} a {period_end.isoformat()}"
                if sales
                else "Sem registros no período selecionado"
            ]
        )
        worksheet.append([])

        dataframe = pd.DataFrame(sales_data)
        if not dataframe.empty:
            worksheet.append(list(dataframe.columns))
            self._style_header_row(worksheet, worksheet.max_row)
            for row in dataframe_to_rows(dataframe, index=False, header=False):
                worksheet.append(row)

        self._auto_adjust_columns(worksheet)
        self._add_summary_sheet(workbook, dataframe, "Vendas")

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        logger.info(
            "sales_excel_export_completed",
            tenant_id=tenant_id,
            record_count=len(sales_data),
        )
        return buffer

    async def export_accuracy_report_to_excel(
        self,
        tenant_id: str,
        start_date: date,
        end_date: date,
    ) -> io.BytesIO:
        """Generate an Excel file with the accuracy report."""

        report_service = self._resolve_report_service()
        report = await report_service.generate_accuracy_report(
            tenant_id=tenant_id,
            start_date=start_date,
            end_date=end_date,
        )

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Resumo Accuracy"

        summary_sheet.append(["Relatório de Accuracy - ConciliaAI"])
        summary_sheet.append([f"Período: {start_date.isoformat()} a {end_date.isoformat()}"])
        summary_sheet.append([])
        summary_sheet.append(["Métrica", "Valor"])
        self._style_header_row(summary_sheet, summary_sheet.max_row)
        summary_sheet.append(["Accuracy Geral", f"{report['overall_accuracy']}%"])
        summary_sheet.append(["Total de Vendas", report["total_sales"]])
        summary_sheet.append(["Matches Realizados", report["total_matches"]])
        summary_sheet.append(["Divergências", report["total_divergences"]])
        self._auto_adjust_columns(summary_sheet)

        trend_sheet = workbook.create_sheet("Evolução Diária")
        trend_data = pd.DataFrame(report.get("trend", []))
        if not trend_data.empty:
            trend_sheet.append(list(trend_data.columns))
            self._style_header_row(trend_sheet, trend_sheet.max_row)
            for row in dataframe_to_rows(trend_data, index=False, header=False):
                trend_sheet.append(row)
        else:
            trend_sheet.append(["Sem dados para o período selecionado"])

        self._auto_adjust_columns(trend_sheet)

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    async def export_divergence_report_to_excel(
        self,
        tenant_id: str,
        start_date: date,
        end_date: date,
    ) -> io.BytesIO:
        """Generate an Excel file with the divergence analysis report."""

        report_service = self._resolve_report_service()
        report = await report_service.generate_divergence_analysis(
            tenant_id=tenant_id,
            start_date=start_date,
            end_date=end_date,
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Análise Divergências"

        sheet.append(["Relatório de Análise de Divergências - ConciliaAI"])
        sheet.append([f"Período: {start_date.isoformat()} a {end_date.isoformat()}"])
        sheet.append([])
        sheet.append(["Métrica", "Valor"])
        self._style_header_row(sheet, sheet.max_row)
        sheet.append(["Total de Divergências", report["total_divergences"]])
        sheet.append(["Valor em Risco (R$)", report["total_amount_at_risk"]])
        sheet.append(["Taxa de Resolução", f"{report['resolution_rate']}%"])
        avg_resolution = report.get("avg_resolution_time_hours")
        sheet.append([
            "Tempo Médio de Resolução (h)",
            avg_resolution if avg_resolution is not None else "N/A",
        ])
        sheet.append([])

        by_type = pd.DataFrame(report.get("by_type", []))
        if not by_type.empty:
            sheet.append(["Divergências por Tipo"])
            sheet.append(list(by_type.columns))
            self._style_header_row(sheet, sheet.max_row)
            for row in dataframe_to_rows(by_type, index=False, header=False):
                sheet.append(row)
            sheet.append([])

        by_severity = pd.DataFrame(report.get("by_severity", []))
        if not by_severity.empty:
            sheet.append(["Divergências por Severidade"])
            sheet.append(list(by_severity.columns))
            self._style_header_row(sheet, sheet.max_row)
            for row in dataframe_to_rows(by_severity, index=False, header=False):
                sheet.append(row)

        self._auto_adjust_columns(sheet)

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def _resolve_report_service(self) -> ReportService:
        if self._report_service is None:
            self._report_service = ReportService(
                sale_repo=self.sale_repo,
                transaction_repo=self.transaction_repo,
                match_repo=self.match_repo,
                divergence_repo=self.divergence_repo,
            )
        return self._report_service

    @staticmethod
    def _serialize_sale(sale: Sale, matched_sales: set[str]) -> dict[str, object]:
        return {
            "ID": sale.id,
            "NSU": str(sale.nsu),
            "Valor (R$)": float(sale.amount.amount),
            "Moeda": sale.amount.currency,
            "Data da Venda": sale.date.isoformat(),
            "Método Pagamento": sale.payment_method,
            "Parcelas": sale.installments,
            "Código Autorização": str(sale.authorization_code)
            if getattr(sale, "authorization_code", None)
            else "",
            "Status": "Reconciliado" if sale.id in matched_sales else "Pendente",
            "Criado em": sale.created_at.isoformat(),
        }

    def _add_summary_sheet(
        self, workbook: Workbook, dataframe: pd.DataFrame, data_type: str
    ) -> None:
        summary = workbook.create_sheet("Resumo")
        summary.append(["Resumo Estatístico"])
        summary.append([f"Tipo: {data_type}"])
        summary.append([f"Total de Registros: {len(dataframe)}"])
        summary.append([])

        if dataframe.empty:
            return

        numeric_columns = dataframe.select_dtypes(include=["number"]).columns
        for column in numeric_columns:
            summary.append([f"Estatísticas - {column}"])
            summary.append(["Métrica", "Valor"])
            self._style_header_row(summary, summary.max_row)

            series = dataframe[column].dropna()
            if series.empty:
                summary.append(["Sem dados numéricos", "-"])
                summary.append([])
                continue

            summary.append(["Média", round(series.mean(), 2)])
            summary.append(["Mediana", round(series.median(), 2)])
            summary.append(["Máximo", round(series.max(), 2)])
            summary.append(["Mínimo", round(series.min(), 2)])
            summary.append([])

        self._auto_adjust_columns(summary)

    def _style_header_row(self, worksheet, row_index: int) -> None:
        for cell in worksheet[row_index]:
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _auto_adjust_columns(worksheet) -> None:
        for column in worksheet.columns:
            values: Iterable[str] = (
                str(cell.value) if cell.value is not None else ""
                for cell in column
            )
            max_length = max((len(value) for value in values), default=0)
            column_letter = column[0].column_letter
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)


__all__ = ["ExportService"]
